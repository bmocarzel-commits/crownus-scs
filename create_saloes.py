from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Mapeamento SCS"

# Colors
header_fill = PatternFill('solid', fgColor='1A1A2E')
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
subheader_fill = PatternFill('solid', fgColor='E8D5B7')
subheader_font = Font(name='Arial', bold=True, color='1A1A2E', size=10)
data_font = Font(name='Arial', size=10, color='333333')
priority_alta = PatternFill('solid', fgColor='FFD6D6')
priority_media = PatternFill('solid', fgColor='FFF3CD')
priority_baixa = PatternFill('solid', fgColor='D4EDDA')
link_font = Font(name='Arial', size=10, color='0066CC', underline='single')
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Title
ws.merge_cells('A1:L1')
ws['A1'] = 'CROWNUS PROFESSIONAL — Mapeamento Estratégico de Salões | São Caetano do Sul'
ws['A1'].font = Font(name='Arial', bold=True, color='FFFFFF', size=14)
ws['A1'].fill = PatternFill('solid', fgColor='0D0D1A')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

# Headers
headers = [
    'Nº', 'SALÃO', 'BAIRRO', 'ENDEREÇO', 'TELEFONE', 'INSTAGRAM',
    'PORTE', 'FOCO PRINCIPAL', 'TRABALHA C/ GLOSS?', 'PRIORIDADE',
    'STATUS ABORDAGEM', 'OBSERVAÇÕES ESTRATÉGICAS'
]
col_widths = [5, 28, 18, 35, 20, 25, 12, 20, 18, 14, 18, 40]

for col, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[2].height = 30

# Salon data
saloes = [
    [1, 'Fathus Cabeleireiros', 'Santo Antônio', 'R. Espírito Santo, 84', '(11) 4221-7887 / 4224-2511', '@fathuscabeleireiros', 'Grande', 'Corte, Coloração, Química', 'Verificar', 'ALTA', 'Não iniciado', 'Desde 1989. Estrutura completa. Potencial alto p/ linha lavatório + tratamento. Salão consolidado = tomador de decisão experiente.'],
    [2, 'Cari Internacional', 'Centro', 'Endereço a confirmar', 'A confirmar', '@cariinternacional', 'Grande', 'Corte, Coloração, Hair Spa', 'Verificar', 'ALTA', 'Não iniciado', 'Um dos mais premiados do ABC. Já trabalha com Hair Spa = entende valor de tratamento. Alvo prioritário.'],
    [3, 'Gate4 Studio', 'Cerâmica', 'Al. Caulim, 115', 'A confirmar', '@gate4studio', 'Médio-Grande', 'Coloração, Transformação', 'Verificar', 'ALTA', 'Não iniciado', '6.8K seguidores. Presença digital forte. Potencial p/ parceria de conteúdo + linha lavatório.'],
    [4, 'Espaço Donna Bellíssima', 'A confirmar', 'A confirmar', 'A confirmar', '@espacodonnabellissima', 'Médio', 'Corte, Coloração, Tricoterapia', 'Verificar', 'ALTA', 'Não iniciado', 'Já oferece TRICOTERAPIA = já entende tratamento capilar. Porta mais fácil p/ Crownus. Prioridade máxima.'],
    [5, 'Studio Mademoiselle', 'A confirmar', 'A confirmar', 'A confirmar', '@studiomademoisellescs', 'Médio', 'Beleza completa', 'Verificar', 'MÉDIA', 'Não iniciado', '3.9K seguidores. Foco em detalhes do look. Avaliar se tem estrutura de lavatório.'],
    [6, 'Sweet Hair São Caetano', 'Santa Paula', 'R. Piauí, 616', 'A confirmar', '@sweethairscs', 'Médio', 'Tratamento, Alisamento', 'Verificar', 'MÉDIA', 'Não iniciado', 'Espaço temático anos 60. Atendimento personalizado. Já trabalha com marca premium (Sweet Hair) = pode ser resistência ou oportunidade.'],
    [7, 'Studio Speciale', 'A confirmar', 'A confirmar', '(11) 95051-4100 / 4228-4990', '@studio_speciale', 'Médio', 'Beleza e Bem-estar', 'Verificar', 'MÉDIA', 'Não iniciado', '4.4K seguidores. Foco em bem-estar = boa entrada p/ cronograma capilar e tratamento.'],
    [8, 'SOU Beleza Integrativa', 'A confirmar', 'A confirmar', 'A confirmar', '@soubelezaintegrativa', 'Médio', 'Beleza Integrativa', 'Verificar', 'MÉDIA', 'Não iniciado', 'Conceito integrativo = alinhado com proposta de tratamento e saúde capilar da Crownus.'],
    [9, 'SOE Centro de Beleza', 'A confirmar', 'A confirmar', '(11) 4221-6099', '@soebrasil', 'Médio-Grande', 'Beleza e Estética Avançada', 'Verificar', 'MÉDIA', 'Não iniciado', '7.9K seguidores. Estética avançada = público que investe. Potencial p/ home care.'],
    [10, 'Camila By Terapia Capilar', 'A confirmar', 'A confirmar', 'A confirmar', '@cambyterapiacapilar', 'Pequeno-Médio', 'Terapia Capilar, Orgânicos', 'Verificar', 'MÉDIA', 'Não iniciado', 'Foco em saúde capilar, ativos naturais. Nicho diferente mas pode abrir porta p/ cronograma capilar.'],
    [11, '1ª Classe Cabeleireiros', 'Centro', 'R. Manoel Coelho, 371', '(11) 4224-6103', 'A buscar', 'Médio', 'Corte, Coloração', 'Verificar', 'MÉDIA', 'Não iniciado', 'Localização central. Funciona até domingo. Alto fluxo = alto consumo de produtos.'],
    [12, 'Mamãe Q Disse', 'A confirmar', 'A confirmar', 'A confirmar', '@salaomamaeqdisse', 'Médio', 'Infantil + Família', 'Verificar', 'BAIXA', 'Não iniciado', '15K seguidores. Foco infantil/família. Não é alvo primário p/ lavatório profissional.'],
    [13, 'Corte Kids São Caetano', 'A confirmar', 'A confirmar', '(11) 4226-4700', '@cortekidssaocaetano', 'Médio', 'Infantil (0-12)', 'Verificar', 'BAIXA', 'Não iniciado', '12K seguidores. Foco 100% infantil. Potencial apenas p/ home care kids se existir.'],
    [14, 'Salão R. Rio de Janeiro', 'Osvaldo Cruz', 'R. Rio de Janeiro, 15', '(11) 4221-9772', 'A buscar', 'A verificar', 'A verificar', 'Verificar', 'A DEFINIR', 'Não iniciado', 'Dados parciais. Necessário visita ou pesquisa adicional.'],
    [15, 'Salão R. João Ramalho', 'Boa Vista', 'R. João Ramalho, 300', '(11) 97053-3442', 'A buscar', 'A verificar', 'Terapia Capilar, Visagismo', 'Verificar', 'MÉDIA', 'Não iniciado', 'Oferece terapia capilar e visagismo = perfil alinhado com Crownus.'],
    [16, 'Salão R. Teodoro Sampaio', 'Cerâmica', 'R. Teodoro Sampaio, 417', '(11) 98286-3467', 'A buscar', 'A verificar', 'A verificar', 'Verificar', 'A DEFINIR', 'Não iniciado', 'Dados parciais. Necessário pesquisa adicional.'],
    [17, 'Salão R. Amazonas', 'Centro', 'R. Amazonas, 750', '(11) 4221-3714', 'A buscar', 'A verificar', 'A verificar', 'Verificar', 'A DEFINIR', 'Não iniciado', 'Dados parciais. Localização central = bom fluxo.'],
    [18, 'Salão R. Conceição', 'Santo Antônio', 'R. Conceição, 732', '(11) 4229-0898', 'A buscar', 'A verificar', 'A verificar', 'Verificar', 'A DEFINIR', 'Não iniciado', 'Dados parciais. Mesmo bairro do Fathus.'],
    [19, 'Spazio Bhio Essenza', 'A confirmar', 'A confirmar', 'A confirmar', '@spaziobhioessenza', 'A verificar', 'A verificar', 'Verificar', 'A DEFINIR', 'Não iniciado', 'Presente no Facebook. Verificar porte e foco.'],
    [20, 'Studio Chico Beauty', 'A confirmar', 'A confirmar', 'A confirmar', '@studiochico', 'A verificar', 'A verificar', 'Verificar', 'A DEFINIR', 'Não iniciado', 'Presente no Booksy com agendamento online. Verificar perfil.'],
]

for row_idx, salon in enumerate(saloes, 3):
    for col_idx, value in enumerate(salon, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = left_wrap if col_idx in [4, 12] else center
        cell.border = thin_border
        
        # Priority colors
        if col_idx == 10:
            if value == 'ALTA':
                cell.fill = priority_alta
                cell.font = Font(name='Arial', bold=True, size=10, color='CC0000')
            elif value == 'MÉDIA':
                cell.fill = priority_media
                cell.font = Font(name='Arial', bold=True, size=10, color='856404')
            elif value == 'BAIXA':
                cell.fill = priority_baixa
                cell.font = Font(name='Arial', bold=True, size=10, color='155724')
        
        # Instagram as link style
        if col_idx == 6 and str(value).startswith('@'):
            cell.font = link_font
    
    ws.row_dimensions[row_idx].height = 45

# Summary sheet
ws2 = wb.create_sheet('Resumo Estratégico')
ws2.merge_cells('A1:D1')
ws2['A1'] = 'RESUMO — MAPEAMENTO SÃO CAETANO DO SUL'
ws2['A1'].font = Font(name='Arial', bold=True, color='FFFFFF', size=13)
ws2['A1'].fill = PatternFill('solid', fgColor='0D0D1A')
ws2['A1'].alignment = center
ws2.row_dimensions[1].height = 35

summary_data = [
    ['MÉTRICA', 'VALOR', '', 'NOTA'],
    ['Total de salões mapeados', 20, '', 'Lista inicial — expandir com campo'],
    ['Prioridade ALTA', 4, '', 'Fathus, Cari, Gate4, Donna Bellíssima'],
    ['Prioridade MÉDIA', 8, '', 'Potencial após validação em campo'],
    ['Prioridade BAIXA', 2, '', 'Foco infantil — baixo fit'],
    ['A DEFINIR', 6, '', 'Dados parciais — completar com visita'],
    ['Já oferecem tratamento capilar', 3, '', 'Donna Bellíssima, Camila By, R. João Ramalho'],
    ['Com forte presença digital (5K+)', 4, '', 'Gate4, SOE, Mamãe Q Disse, Corte Kids'],
    ['', '', '', ''],
    ['PRÓXIMOS PASSOS', '', '', ''],
    ['1', 'Completar dados faltantes (endereço, Instagram, porte)', '', 'Visita de campo ou pesquisa'],
    ['2', 'Verificar quais trabalham com Gloss', '', 'Priorizar esses na abordagem'],
    ['3', 'Iniciar engajamento no Instagram dos ALTA', '', 'Semana 1-2 da estratégia'],
    ['4', 'Preparar material de abordagem com ângulo "slot vazio"', '', 'Landing page + convite evento'],
    ['5', 'Abordagem presencial com convite do evento', '', 'Começar pelos 4 ALTA'],
]

for row_idx, row_data in enumerate(summary_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = left_wrap
        cell.border = thin_border
        if row_idx == 3:
            cell.font = subheader_font
            cell.fill = subheader_fill
        if row_idx == 12:
            cell.font = Font(name='Arial', bold=True, size=10, color='1A1A2E')
            cell.fill = PatternFill('solid', fgColor='E8D5B7')

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 50
ws2.column_dimensions['C'].width = 5
ws2.column_dimensions['D'].width = 50

# Freeze panes
ws.freeze_panes = 'A3'
ws2.freeze_panes = 'A3'

# Auto filter
ws.auto_filter.ref = f'A2:L{len(saloes)+2}'

output = '/sessions/great-exciting-newton/mnt/outputs/Crownus_Mapeamento_SCS.xlsx'
wb.save(output)
print(f'Saved to {output}')
